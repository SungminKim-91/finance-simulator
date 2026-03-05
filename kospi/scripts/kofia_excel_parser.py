#!/usr/bin/env python3
"""
KOFIA FreeSIS 엑셀 파싱 — "한눈에 보는 자본시장통계" 엑셀 → timeseries 필드 변환.

엑셀 구조 (56행×5열, 시트명 "한눈에 보는 자본시장통계"):
  A열: 지표명 (계층 `>` 구분, 줄바꿈 포함)
  B열: 기준일자 (YY/MM/DD 문자열)
  C열: 값 (천 단위 쉼표 문자열)
  D열: 증감, E열: 증감율(%)

Usage:
    from scripts.kofia_excel_parser import parse_kofia_excel
    result = parse_kofia_excel("path/to/file.xlsx")
    # → {"2026-03-04": {"deposit_billion": 132068.238, ...}, ...}
"""
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

# 엑셀 A열 지표명 → (timeseries 필드, 나눗수)
# 나눗수: 백만원→십억원 = /1000, 억원→십억원 = /10, None = 그대로
FIELD_MAP = {
    "자금동향>증시자금>투자자예탁금": ("deposit_billion", 1e3),
    "자금동향>증시자금>위탁매매미수금": ("unsettled_billion", 1e3),
    "자금동향>증시자금>실제반대매매": ("forced_liq_billion", 1e3),
    "자금동향>신용공여>신용거래융자": ("credit_balance_billion", 1e3),
    "주식시장>국내지수>KOSPI": ("kospi", None),
    "주식시장>거래대금>KOSPI": ("kospi_trading_value_billion", 10),
}


def _normalize_label(raw: str) -> str:
    """A열 지표명 정규화: 줄바꿈 제거, 공백 trim."""
    if not raw:
        return ""
    return raw.replace("\n", "").replace("\r", "").strip()


def _parse_date(raw) -> str | None:
    """B열 날짜 변환: 'YY/MM/DD' → 'YYYY-MM-DD'."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    parts = s.split("/")
    if len(parts) == 3:
        yy, mm, dd = parts
        year = int(yy)
        if year < 100:
            year += 2000
        return f"{year:04d}-{int(mm):02d}-{int(dd):02d}"
    return None


def _parse_value(raw) -> float | None:
    """C열 값 변환: 쉼표 제거 → float."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_kofia_excel(path: str | Path) -> dict[str, dict]:
    """엑셀 파싱 → {date: {field: value}} 반환.

    날짜별 그룹핑 (지표마다 기준일자가 다를 수 있음).
    """
    if openpyxl is None:
        raise ImportError("openpyxl 필요: pip install openpyxl")

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # 시트 찾기: 정확한 이름 또는 첫 번째 시트
    sheet_name = "한눈에 보는 자본시장통계"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    result: dict[str, dict] = {}
    matched = 0

    for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
        if not row or not row[0]:
            continue

        label = _normalize_label(str(row[0]))
        if label not in FIELD_MAP:
            continue

        field_name, divisor = FIELD_MAP[label]
        date_str = _parse_date(row[1])
        raw_val = _parse_value(row[2])

        if date_str is None or raw_val is None:
            continue

        # 단위 변환
        if divisor is not None:
            value = round(raw_val / divisor, 3)
        else:
            value = round(raw_val, 2)

        if date_str not in result:
            result[date_str] = {}
        result[date_str][field_name] = value
        matched += 1

    wb.close()
    print(f"  [EXCEL] {path.name}: {matched}개 필드 파싱, {len(result)}개 날짜")
    return result
